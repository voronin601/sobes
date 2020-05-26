from django.shortcuts import render
import requests as re
import json
from .models import data
import datetime
from django.http import JsonResponse
from django.http import HttpResponse
import os
from openpyxl import Workbook

###Скачиваение данных, заполнение бд и json/xlsx
'''def download_json(requests):
    a = re.get('https://hubofdata.ru/storage/f/2013-08-18T19%3A58%3A51.196Z/greetings-data.json')
    for idd, i in enumerate(json.loads(a.text)['items']):
        data(category = i['category'], from_f = i['from'], title = i['title'], text = i['text'], date = dat(i['thedate']), idd = i['id']).save()
        print(idd, end = '\r')

def dat(date):
    d = ''
    a = ['янв', 'фев', 'мар', 'апре', 'мая', 'июн', 'июл', 'авгу', 'сент', 'октя', 'нояб', 'декаб']
    for idd, i in enumerate(a):
        if i in date: d = str(idd + 1)
    return date.split(' ')[2] + '-' + d + '-' + date.split(' ')[0]

def download_xlsx(requests):
    a = re.get('https://hubofdata.ru/storage/f/2013-08-18T19%3A58%3A51.196Z/greetings-data.json')
    b = json.loads(a.text)['items']
    wb = Workbook()
    ws = wb.active
    st = ['A', 'B', 'C', 'D', 'E', 'F']
    for idd, i in enumerate(data._meta.fields):
        ws['%s1' %st[idd]] = i.name
    for idd, i in enumerate(b):
        print(idd)
        for idd_1, j in enumerate(b[0].keys()):
            ws['%s%s' %(st[idd_1], str(idd + 2))] = i[j]
    wb.save('dat.xlsx')'''

#возвращает весь json файл
def ret_json(requests):
    with open('./first/templates/data.json', 'r') as t:
        a = json.loads(t.read())
    return JsonResponse(a, safe = False)

#возвращает весь xlsx файл
def ret_xlsx(requests):
    if os.path.exists(os.getcwd() + '\\first\\templates\\dat.xlsx'):
        with open('./first/templates/dat.xlsx', 'rb') as fl:
            response = HttpResponse(fl.read(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=data.xlsx'
            return response

    return HttpResponse(requests, "Error")

def ret_json_date(requests, date):
    try:
        post = data.objects.filter(date = datetime.datetime.strptime(date,'%Y-%m-%d').date())
        ret = [{"category":i.category, "from":i.from_f, "title":i.title, "text":i.text, "thedate":date, "id":i.id} for i in post]
        return JsonResponse(ret, safe = False)
    except:
        return JsonResponse({'error':'No such date'}, safe = False)

def ret_xlsx_date(requests, date):
    try:
        post = data.objects.filter(date = datetime.datetime.strptime(date,'%Y-%m-%d').date())
        wb = Workbook()
        ws = wb.active
        st = ['A', 'B', 'C', 'D', 'E', 'F']
        for idd, i in enumerate(data._meta.fields):
            if idd != 0:
                ws['%s1' %st[idd-1]] = i.name
        for idd, i in enumerate(post):
            ws['A%s' %(str(idd + 2))] = i.category
            ws['B%s' %(str(idd + 2))] = i.from_f
            ws['C%s' %(str(idd + 2))] = i.title
            ws['D%s' %(str(idd + 2))] = i.text
            ws['E%s' %(str(idd + 2))] = i.date
            ws['F%s' %(str(idd + 2))] = i.idd
        wb.save('date.xlsx')
        with open('date.xlsx', 'rb') as fl:
            response = HttpResponse(fl.read(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=data_%s.xlsx' %date
            return response
    except:
        return JsonResponse({'error':'No such date'}, safe = False)